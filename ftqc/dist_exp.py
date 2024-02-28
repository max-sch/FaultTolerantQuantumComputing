import pandas as pd
from provider.circuit_provider import QasmBasedCircuitProvider
from core.qerror_detection import MeasurementNoiseQuantifier
from core.entities import Measurements
from core.entities import QuantumComputerSimulator
from qiskit.tools.visualization import plot_histogram

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
import os

def evaluateDistances():
    quantifiers = {
    "Shannon Entropy": MeasurementNoiseQuantifier.using_shannon_entropy(0.1),
    "KL Divergence": MeasurementNoiseQuantifier.using_kl_divergence(0.1),
    "Cross Entropy": MeasurementNoiseQuantifier.using_cross_entropy(0.1),
    "Jensen-Shannon Divergence": MeasurementNoiseQuantifier.using_jensen_shannon_divergence(0.1),
    "Bhattacharyya": MeasurementNoiseQuantifier.using_bhattacharyya(0.1),
    "Hellinger": MeasurementNoiseQuantifier.using_hellinger(0.1)
    }

    # Evaluate measurements using the provided distances
    evaluation_results = []
    qasm_files_dir = "MQTBench"
    provider = QasmBasedCircuitProvider(dir=qasm_files_dir)  

    image_files = []
    if not os.path.exists('histograms'):
        os.mkdir('histograms')

    for batch in provider.get():  
        for noise in [False, True]:
            if noise:
                simulator = QuantumComputerSimulator.create_noisy_simulator() 
                simulator.modify_noise()

            else:
                simulator = QuantumComputerSimulator.create_perfect_simulator()
            
            results = simulator.execute_batch(batch.circuits)
            measurements = []
            for circuit in batch.circuits:
                
                counts = results.get_counts(circuit.qiskit_circuit)
                measurement = Measurements("channel", counts)
                

                image_filename = os.path.join("histograms", f"{circuit.id}-noise:{noise}.png")
                plot_histogram(counts, filename=image_filename, bar_labels=False)
                image_files.append(image_filename)

                results_for_this_circuit = {
                "circuit_name": circuit.id,
                "Noise": noise,
                "count": "counts"
                }
                
                for name, quantifier in quantifiers.items():
                    results_for_this_circuit[name] = quantifier.measure_closeness_to_uniform_dist(measurement)
                    
                evaluation_results.append(results_for_this_circuit) 

    df_evaluation = pd.DataFrame(evaluation_results)
    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df_evaluation, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for idx, image_file in enumerate(image_files, 1):
        img = Image(image_file)
       
        ws.column_dimensions['C'].width = img.width // 6  
        image_col_width = ws.column_dimensions['C'].width
        ws.row_dimensions[idx + 1].height = img.height 
        ws.add_image(img, f"C{idx+1}")

    font = Font(bold=True, size=40)
    alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        col_dimension = ws.column_dimensions[col[0].column_letter]
        if col_dimension.width < image_col_width:
            col_dimension.width = image_col_width
  
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment


    # Save the workbook
    wb.save("results.xlsx")

def evaluateCorrelation():
    quantifiers = {
    "Shannon Entropy": MeasurementNoiseQuantifier.using_shannon_entropy(0.1),
    "KL Divergence": MeasurementNoiseQuantifier.using_kl_divergence(0.1),
    "Cross Entropy": MeasurementNoiseQuantifier.using_cross_entropy(0.1),
    "Jensen-Shannon Divergence": MeasurementNoiseQuantifier.using_jensen_shannon_divergence(0.1),
    "Bhattacharyya": MeasurementNoiseQuantifier.using_bhattacharyya(0.1),
    "Hellinger": MeasurementNoiseQuantifier.using_hellinger(0.1)
    }

    # Evaluate measurements using the provided distances
    evaluation_results = []
    qasm_files_dir = "MQTBench"
    provider = QasmBasedCircuitProvider(dir=qasm_files_dir)  

    image_files = []
    if not os.path.exists('histograms'):
        os.mkdir('histograms')

    for batch in provider.get():          
        simulator = QuantumComputerSimulator.create_noisy_simulator() 
        simulator.modify_noise()

        
        perfect_simulator = QuantumComputerSimulator.create_perfect_simulator()
        
        results = simulator.execute_batch(batch.circuits)
        perfect_results = perfect_simulator.execute_batch(batch.circuits)
        for circuit in batch.circuits:
            
            counts = results.get_counts(circuit.qiskit_circuit)
            perfect_counts = perfect_results.get_counts(circuit.qiskit_circuit)

            measurement = Measurements("channel", counts)
            greatest_rank = max(counts, key=counts.get)
            perfect_rank = max(perfect_counts, key=counts.get)
            
            if (greatest_rank == perfect_rank ):
                match = True
            else:
                match = False

            image_filename = os.path.join("histograms", f"{circuit.id}.png")
            plot_histogram(counts, filename=image_filename, bar_labels=False)
            image_files.append(image_filename)

            results_for_this_circuit = {
            "circuit_name": circuit.id,
            "count": "counts",
            "match": match
            }
            
            for name, quantifier in quantifiers.items():
                results_for_this_circuit[name] = quantifier.measure_closeness_to_uniform_dist(measurement)
                
            evaluation_results.append(results_for_this_circuit)

    df_evaluation = pd.DataFrame(evaluation_results)
    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df_evaluation, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for idx, image_file in enumerate(image_files, 1):
        img = Image(image_file)
       
        ws.column_dimensions['J'].width = img.width // 6  
        image_col_width = ws.column_dimensions['J'].width
        ws.row_dimensions[idx + 1].height = img.height 
        ws.add_image(img, f"J{idx+1}")

    correlation_results = {"Metric": [], "Correlation with Match": []}
    for quantifier_name in quantifiers.keys():
        correlation = df_evaluation["match"].astype(int).corr(df_evaluation[quantifier_name])
        correlation_results["Metric"].append(quantifier_name)
        correlation_results["Correlation with Match"].append(correlation)

    # Convert the correlation results in a DataFrame and insert into workbook
    correlation_df = pd.DataFrame(correlation_results)
    for r_idx, row in enumerate(dataframe_to_rows(correlation_df, index=False, header=True), ws.max_row + 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    font = Font(bold=True, size=40)
    alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        col_dimension = ws.column_dimensions[col[0].column_letter]
        if col_dimension.width < image_col_width:
            col_dimension.width = image_col_width

  
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment


    # Save the workbook
    wb.save("results.xlsx")

if __name__ == '__main__':
    evaluateCorrelation()

    



